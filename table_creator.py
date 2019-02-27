# -*- coding: utf-8 -*-
"""
Created on Wed Feb 13 17:51:54 2019

@author: Draguve
         Rishi
         Sarthak

"""
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles.borders import Border, Side

import sqlite3
from openpyxl.styles import PatternFill
# from openpyxl.compat import range as pyxlrange
from copy import copy
import colorsys

import traceback


def find_all_batches(ws,class_row):
    # class_row = None
    # for row in ws.iter_rows(min_row=1, max_col=1, max_row=10):
    #     for cell in row:
    #         if cell.value is not None and cell.value.lower() == "day":
    #             class_row = cell.row + 1
    # if class_row is None:
    #     return None
    class_row = ws[class_row].row
    classes = []
    for column in ws.iter_cols(min_row=class_row, max_col=100, max_row=class_row):
        for cell in column:
            if cell.value is not None:
                if str(cell.value).lower() != "day" and str(cell.value).lower() != "hours":
                    classes.append(cell.value)
    return classes


def find_batch(ws, class_code,class_row):
    # class_row = None
    # for row in ws.iter_rows(min_row=1, max_col=1, max_row=10):
    #     for cell in row:
    #         if cell.value is not None and cell.value.lower() == "day":
    #             class_row = cell.row + 1
    # if class_row is None:
    #     return None
    class_row = ws[class_row].row
    for column in ws.iter_cols(min_row=class_row, max_col=100, max_row=class_row):
        for cell in column:
            if str(cell.value) is not None and str(cell.value).lower() == str(class_code).lower():
                return cell
    return None


def get_timetable(ws, name_cell):
    finalworkbook, finalsheet = create_empty_table()
    finalsheet.title = name_cell.value

    merge_dict = get_merge_dict(ws)

    to_skip = 0
    start_cell = name_cell.offset(1, 0)

    daycell = finalsheet["B2"]
    current_cell = daycell

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for x in range(5):
        for row in ws.iter_rows(min_row=start_cell.row, max_row=(int(start_cell.row) + 19),
                                max_col=column_index_from_string(start_cell.column),
                                min_col=column_index_from_string(start_cell.column)):
            for cell in row:
                if to_skip <= 0:
                    # get the period data at the right location
                    class_code, class_room, to_skip = get_period(merge_dict, cell)
                    if class_code is not None and class_code[-1] != "P" and to_skip == 3:
                        print(cell)
                    if class_code is not None and class_room is not None:
                        to_write = current_cell
                        to_write.value = class_code
                        if to_skip == 3:
                            to_write.offset(1, 0).value = class_room
                            to_write.offset(2, 0).value = "LAB"
                            # style
                            style_range(finalsheet, '{0}{1}:{2}{3}'.format(to_write.column, to_write.row,
                                                                           to_write.offset(3, 0).column,
                                                                           to_write.offset(3, 0).row), border=border)
                        else:
                            to_write.offset(1, 0).value = "%s" % (class_room)
                            # style
                            style_range(finalsheet, '{0}{1}:{2}{3}'.format(to_write.column, to_write.row,
                                                                           to_write.offset(1, 0).column,
                                                                           to_write.offset(1, 0).row), border=border)
                    current_cell = current_cell.offset(to_skip + 1, 0)
                else:
                    to_skip -= 1
                end_cell = cell
        # End of day
        to_skip = 0
        start_cell = end_cell.offset(1, 0)
        daycell = daycell.offset(0, 1)
        current_cell = daycell
    return finalworkbook


def create_table(ws, merge_dict, class_cell, table_name):
    start_cell = class_cell.offset(1, 0)
    conn = sqlite3.connect('time_tables.db')
    to_skip = 0
    conn.execute('''CREATE TABLE "{}"(
                    DAY INT,
                    START_TIME INT,
                    END_TIME INT,
                    TYPE TEXT,
                    DATA TEXT,
                    CLASS_CODE TEXT
                    )'''.format(table_name))
    for x in range(5):
        start_time = 8
        for row in ws.iter_rows(min_row=start_cell.row, max_row=(int(start_cell.row) + 19),
                                max_col=column_index_from_string(start_cell.column),
                                min_col=column_index_from_string(start_cell.column)):
            for cell in row:
                if to_skip <= 0:
                    # get the period data at the right location
                    data, to_skip, class_code = get_period_data(merge_dict, cell)

                    if to_skip == 3:
                        end_time = start_time + 2
                    else:
                        end_time = start_time + 1

                    if data is not None:
                        y = 'INSERT INTO "{}" VALUES({},{},{},"{}","{}","{}")'.format(table_name, x, start_time, end_time,
                                                                               class_code[-1], data,class_code)
                        print(y)
                        conn.execute(y)
                else:
                    to_skip -= 1
                end_cell = cell
                start_time = end_time
        # End of day
        to_skip = 0
        start_cell = end_cell.offset(1, 0)
    conn.commit()
    conn.close()


def get_period_data(merged, cell):
    try:
        class_cell = merged[cell]
    except:
        return None, 1 , None
    class_code = class_cell.value

    if class_code is not None and class_code[-1] == "P":
        data = str(class_cell.value) + ':' + str(class_cell.offset(1, 0).value) + ':' + str(
            class_cell.offset(2, 0).value) + ':' + str(class_cell.offset(3, 0).value)
        to_skip = 3
    elif class_code is not None:
        data = str(class_cell.value) + ':' + str(class_cell.offset(1, 0).value)
        to_skip = 1
    else:
        return None,1,None
    return data, to_skip , class_code


def get_period(merged, cell):
    try:
        class_cell = merged[cell]
    except:
        return None, None, 1
    class_code = class_cell.value

    # find cells to skip
    if class_code is not None and class_code[-1] == "P":
        to_skip = 3
    else:
        to_skip = 1

    # find class room
    if class_code is not None:
        class_room = class_cell.offset(1, 0).value
    else:
        class_room = None

    return class_code, class_room, to_skip


def get_merge_dict(sheet):
    ranges = sheet.merged_cells.ranges
    final = {}
    for mergedcell in ranges:
        bounds = mergedcell.bounds
        for i in range(bounds[0], bounds[2] + 1):
            for j in range(bounds[1], bounds[3] + 1):
                final[sheet.cell(j, i)] = sheet.cell(bounds[1], bounds[0])
    return final


def create_empty_table():
    # creating a new workbook to store the new timetable
    wb = Workbook()
    finalsheet = wb.active

    # formatting the table
    finalsheet["A2"].value = "Time/Day"

    current_cell = finalsheet["A2"]
    time = 8
    for x in range(1, 12):
        current_cell.value = str(time % 12) + " To"
        current_cell.offset(1, 0).value = str((time + 1) % 12)
        current_cell = current_cell.offset(2, 0)
        time += 1

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    current_cell = finalsheet["B1"]
    for x in range(5):
        current_cell.value = days[x]
        current_cell = current_cell.offset(0, 1)

    return wb, finalsheet


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def ask_question(question, choices):
    while True:
        print(question)
        for x in range(0, len(choices)):
            print("{0}) - {1}".format(x + 1, choices[x]))
        response = input(">")
        try:
            if int(response) < len(choices) + 1:
                return int(response) - 1
        except ValueError:
            print("Please enter a valid choice")


if __name__ == "__main__":
    # Load xlsx file
    while True:
        try:
            print("Please input filename for the timetable")
            response = input(">")
            wb = load_workbook(response)
            break
        except FileNotFoundError:
            print("please check the filename provided")

    for sheet in wb.sheetnames:
        worksheet = wb[sheet]
        while True:
            try:
                print("Please input the first batches cell for worksheet {}".format(sheet))
                class_row = input('>')
                break
            except ValueError:
                print("Input Valid answer")
        batches = find_all_batches(worksheet,class_row)
        merge_dict = get_merge_dict(worksheet)
        for batch in batches:
            try:
                batch_cell = find_batch(worksheet,batch,class_row)                
                create_table(worksheet,merge_dict,batch_cell,"t"+sheet.replace(',','').replace(' ','')+"_"+batch_cell.value.replace(' ',''))
            except Exception as e:
                print(batch_cell.value)

